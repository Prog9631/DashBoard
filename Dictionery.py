import openpyxl
wb = openpyxl.load_workbook('data_sheet.xlsx')
sheet = wb['Sheet3']
Words = {"Production", "production"}

for colNum in range(1, sheet.max_column):
    stringinexcel = sheet.cell(row=1, column=colNum).value
    #print(stringinexcel)
    if stringinexcel in Words:
            print("yes")
        


